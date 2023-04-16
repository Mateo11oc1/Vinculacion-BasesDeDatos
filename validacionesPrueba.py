import openpyxl
import pandas
import math

"""Devuelve la columna como un diccionario de acuerdo a los parametros"""
def leerColumna(leido,  columna: int):
    
    lista = leido.iloc[7:,columna].tolist()
    #[7:,columna] nos da la celda desde la fila 7 hasta la ultima fila,  columna 9 y le hace lista
    columna = {"atractor": lista[0], "numAtractores": lista[2], "tamanio": lista[3:6], "jornada": lista[6:11], "dias": lista[12:22], "numColumna": columna, "hoja": 9}
    return columna

"""validar que la suma de atractores no sea mayor o 
menor al numero de atractores colocado"""
def validarSuma(columna):
    try:
        #si es que hay algo en la celda se valida que el numero de atractores sea igual a la suma de los tamanos
        if columna["numAtractores"] == sum(x for x in columna["tamanio"] if not math.isnan(x)):
            return True
        else:
            return False
    except TypeError:
        #la excepecion se da si el dato no es un numero(error de tipos)
        return False
    
"""Validar que la suma de todos los datos en días de atención no sea menor al numero de atractores"""
def validarSumaDias(columna:dict):
    #si el numero de atractores es mayor a la suma de los valores de dias
    if columna["numAtractores"] > sum(x for x in columna['dias'] if not math.isnan(x)):
        return False
    else:
        return True
"""Validar que la suma de todos los datos en jornada no sea menor al numero de atractores"""
def validarSumaJornada(columna:dict):

    #si el numero de atractores es mayor a la suma de los valores del jornada
    if columna["numAtractores"] > sum(x for x in columna['jornada'] if not math.isnan(x)):
        return False
    else:
        return True

"""Validar que uno o varios de los datos de la jornada no sobrepase el numero de atractores"""
"""No se valida que esto se cumpla con la suma de los datos de la jornada debido a que se puede tener un atractor
que sea matutino y nocturno a la vez"""
def validarJornadaNoSobrepaseAtractores(columna:dict):
    for i in range(len(columna['jornada'])):
        if columna["numAtractores"] < columna['jornada'][i]:
            return False

    return True


"""Validar que uno o varios de los datos de los días de atención no sobrepasen el numero de atractores"""
"""No se valida que esto se cumpla con la suma de los datos de los días de atención debido a que se puede tener un atractor
que atienda lunes y martes a la vez"""
def validarDiasNoSobrepaseAtractores(columna:dict):
    for i in range(len(columna['dias'])):
        if columna["numAtractores"] < columna['dias'][i]:
            return False

    return True

"""Validar que solo sean numero y no letras"""
def validarCaracteres(columna: dict):
    #si no es un numero entero
    if not isinstance(columna["numAtractores"], int):
        return False
    #se recorre la lista de valores de la columna desde 2 en adelante porque va desde el tamanio
    for i in list(columna.values())[2:]:
        #se recorre cada lista, porque hay la lista tamanio, jornada, dias
        for j in i:
            #si el valor de la celda es un string o esta vacio o es un numero decimal
            if isinstance(j, str) or (not math.isnan(j) and not isinstance(j, int)):
                return False
    
    return True


#sirve para validar que no se ingresen numeros muy grandes o muy pequenios o de punto flotante
#incluye a los numeros negativos
def validarExtremos(columna:dict):
    for i in list(columna.values())[2:]:
        for j in i:
            if not math.isnan(j):
                if j>1000 or j<1 or isinstance(j,float):
                    return False
    return True

"""Validar que en caso de haber datos en las filas inferiores
, el campo de numero de atractores no sea nulo"""
def validarNaN(columna: dict):
#  Valida que el detalle, las filas desde el tamanio hacaia abajo no contengan datos

    def validarDetalle(columna:dict):
        #recorre la columna desde la fila 2 hacia abajo
        for i in list(columna.values())[2:]:
            for j in i:
                #si el valor de la celda no es entero, retorna falso
                if not isinstance(j, int):
                    return False
                #si no esta vacio retorna falso
                elif not math.isnan(j):
                    return False
                else: 
                    return True

    #si el numero de atractores es nulo y el detalle no esta vacio
    if math.isnan(columna["numAtractores"]) and not validarDetalle(columna):
        return False
    else:
        return True
    

"""En caso de que el numero de atractores sea nulo, y hayan datos en las columnas de abajo
corregir sumando los datos del apartado tamanio"""
def corregirAtractoresNulos(columna: dict):
    def verificarDatos(columna: dict):
        for i in columna["tamanio"]:
            pass

leido = pandas.read_excel("../04. Forumularios digitalizados grupo 4.xlsx", sheet_name=9)
columna = leerColumna(leido, 9)

#se leera el excel con una libreria que permite escribir en el archivo .xlsx
workbook = openpyxl.load_workbook("../04. Forumularios digitalizados grupo 4.xlsx")

"""print(validarSuma(columna))
print(validarCaracteres(columna))
print(validarNaN(columna))
print(validarExtremos(columna))
print(validarSumaJornada(columna))
print(validarSumaDias(columna))
print(validarJornadaNoSobrepaseAtractores(columna))
print(validarDiasNoSobrepaseAtractores(columna))"""

"""metodo que sirve si es que se indica que hay 3 atractores de un tipo x grandes y 1 mediano y no se indica el 
numero de atractores totales, entonces se modifica el archivo desde el codigo en numero de atractores"""
def modificarCampoNAtractores(columna: dict):
    #si el numero de atractores es nulo
    print("hola pata")
    if math.isnan(columna["numAtractores"]) :
        hoja=workbook.worksheets[9]
        print(sum(x for x in columna['tamanio'] if not math.isnan(x)))
        hoja.cell(row=11, column=10).value=sum(x for x in columna['tamanio'] if not math.isnan(x))
        workbook.save("../04. Forumularios digitalizados grupo 4.xlsx")
    else:
        return True

"""metodo para ver si esta marcado matutino y vespertino es diurno"""

def corregirDiurno(columna: dict):

    if not math.isnan(sum(x for x in columna["jornada"] if not math.isnan(x))):
        
        if not math.isnan(columna["jornada"][0]) and not math.isnan(columna["jornada"][1]) and columna["jornada"][0] == columna["numAtractores"] and columna["jornada"][1] == columna["numAtractores"]:
            hoja = workbook.worksheets[columna["hoja"]]
            hoja.cell(row = 17, column = columna["numColumna"] + 1).value = columna["numAtractores"]
            hoja.cell(row = 15, column = columna["numColumna"] + 1).value = ""
            hoja.cell(row = 16, column = columna["numColumna"] + 1).value = ""
            workbook.save("../04. Forumularios digitalizados grupo 4.xlsx")



modificarCampoNAtractores(columna)
corregirDiurno(columna)
